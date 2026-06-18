# helpers/excel_generator.py
"""
Kertas Kerja Evaluasi (.xlsx) berbasis kriteria, dengan rumus hidup & link antar-sheet.

Struktur:
- 'Peserta'           : data penyedia + harga penawaran/terkoreksi.
- 'Evaluasi Teknis'   : baris = kriteria TEKNIS dari LKE (berbobot + ambang batas),
                        kolom = peserta. Nilai tertimbang dihitung SUMPRODUCT(bobot, nilai),
                        lulus bila >= ambang batas teknis.
- 'Evaluasi Kualifikasi': baris = persyaratan dari LDK, kolom = peserta (Memenuhi/Tidak),
                        lulus bila tidak ada "Tidak Memenuhi" (COUNTIF).
- 'Rekapitulasi'      : menarik hasil administrasi/teknis/kualifikasi + harga via rumus,
                        menentukan hasil akhir, peringkat harga, dan usulan pemenang.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from core.config import Config

_HF = Font(bold=True, color="FFFFFF")
_FILL = PatternFill("solid", fgColor="11515F")
_SUB = PatternFill("solid", fgColor="D5E8F0")
_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RP = '#,##0'


def _hdr(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HF; cell.fill = _FILL; cell.alignment = _CTR


def generate_xlsx_kertas_kerja(paket, peserta_list, kriteria_list):
    wb = Workbook()
    n = len(peserta_list)
    teknis = [k for k in kriteria_list if k.kelompok == 'Teknis']
    kualif = [k for k in kriteria_list if k.kelompok == 'Kualifikasi']

    # ---------------- Sheet Peserta ----------------
    ws = wb.active; ws.title = "Peserta"
    ws["A1"] = f"KERTAS KERJA EVALUASI — {paket.nama_paket}"; ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Metode: {paket.metode_pemilihan} | HPS: {paket.nilai_hps:,.0f}"
    for i, h in enumerate(["No", "Nama Penyedia", "NPWP", "Harga Penawaran", "Harga Terkoreksi"], 1):
        ws.cell(row=4, column=i, value=h)
    _hdr(ws, 4, 5)
    for idx, ps in enumerate(peserta_list):
        r = 5 + idx
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=ps.nama_penyedia)
        ws.cell(row=r, column=3, value=ps.npwp or "")
        ws.cell(row=r, column=4, value=float(ps.harga_penawaran or 0)).number_format = _RP
        ws.cell(row=r, column=5, value=float(ps.harga_terkoreksi or 0)).number_format = _RP
    ws.column_dimensions["B"].width = 30

    pcol0 = 5  # kolom peserta pertama di sheet teknis (E)

    # ---------------- Sheet Evaluasi Teknis (dari LKE) ----------------
    wt = wb.create_sheet("Evaluasi Teknis")
    wt["A1"] = "EVALUASI TEKNIS (disusun dari LKE pada LDP)"; wt["A1"].font = Font(bold=True, size=12)
    head = ["No", "Unsur/Kriteria Teknis (LKE)", "Bobot (%)", "Ambang"] + [ps.nama_penyedia for ps in peserta_list]
    for i, h in enumerate(head, 1):
        wt.cell(row=3, column=i, value=h)
    _hdr(wt, 3, len(head))
    first = 4
    for j, k in enumerate(teknis):
        r = first + j
        wt.cell(row=r, column=1, value=j + 1)
        wt.cell(row=r, column=2, value=k.nama_kriteria)
        wt.cell(row=r, column=3, value=float(k.bobot or 0))
        wt.cell(row=r, column=4, value=float(k.ambang_batas or 0))
        for i in range(n):
            wt.cell(row=r, column=pcol0 + i, value=None)  # diisi Pokja (0-100)
    last = first + len(teknis) - 1 if teknis else first
    bobot_rng = f"$C${first}:$C${last}"
    # baris hasil
    r_nilai = last + 1
    r_ambang = last + 2
    r_hasil = last + 3
    wt.cell(row=r_nilai, column=2, value="NILAI TERTIMBANG").font = Font(bold=True)
    wt.cell(row=r_ambang, column=2, value="Ambang Batas Teknis (nilai minimal)").font = Font(bold=True)
    wt.cell(row=r_ambang, column=4, value=70)  # default, dapat diubah Pokja
    wt.cell(row=r_hasil, column=2, value="HASIL TEKNIS").font = Font(bold=True)
    for i in range(n):
        col = get_column_letter(pcol0 + i)
        if teknis:
            wt.cell(row=r_nilai, column=pcol0 + i,
                    value=f"=SUMPRODUCT({bobot_rng},{col}{first}:{col}{last})/100")
            wt.cell(row=r_hasil, column=pcol0 + i,
                    value=f'=IF({col}{r_nilai}>=$D${r_ambang},"LULUS","GUGUR")')
        else:
            wt.cell(row=r_hasil, column=pcol0 + i, value='="(tanpa kriteria teknis)"')
    wt.column_dimensions["B"].width = 42

    # ---------------- Sheet Evaluasi Kualifikasi (dari LDK) ----------------
    wk = wb.create_sheet("Evaluasi Kualifikasi")
    wk["A1"] = "EVALUASI KUALIFIKASI (disusun dari persyaratan LDK)"; wk["A1"].font = Font(bold=True, size=12)
    head = ["No", "Persyaratan Kualifikasi (LDK)"] + [ps.nama_penyedia for ps in peserta_list]
    for i, h in enumerate(head, 1):
        wk.cell(row=3, column=i, value=h)
    _hdr(wk, 3, len(head))
    kcol0 = 3
    kfirst = 4
    dv = DataValidation(type="list", formula1='"Memenuhi,Tidak Memenuhi"', allow_blank=True)
    wk.add_data_validation(dv)
    for j, k in enumerate(kualif):
        r = kfirst + j
        wk.cell(row=r, column=1, value=j + 1)
        wk.cell(row=r, column=2, value=k.nama_kriteria)
        for i in range(n):
            dv.add(wk.cell(row=r, column=kcol0 + i))
    klast = kfirst + len(kualif) - 1 if kualif else kfirst
    rk_hasil = klast + 1
    wk.cell(row=rk_hasil, column=2, value="HASIL KUALIFIKASI").font = Font(bold=True)
    for i in range(n):
        col = get_column_letter(kcol0 + i)
        if kualif:
            wk.cell(row=rk_hasil, column=kcol0 + i,
                    value=f'=IF(COUNTIF({col}{kfirst}:{col}{klast},"Tidak Memenuhi")=0,"LULUS","GUGUR")')
        else:
            wk.cell(row=rk_hasil, column=kcol0 + i, value='="(tanpa persyaratan)"')
    wk.column_dimensions["B"].width = 46

    # ---------------- Sheet Rekapitulasi ----------------
    wr = wb.create_sheet("Rekapitulasi")
    wr["A1"] = "REKAPITULASI HASIL EVALUASI"; wr["A1"].font = Font(bold=True, size=12)
    head = ["No", "Penyedia", "Administrasi", "Teknis", "Kualifikasi",
            "Harga Terkoreksi", "Hasil Akhir", "Peringkat", "Usulan"]
    for i, h in enumerate(head, 1):
        wr.cell(row=3, column=i, value=h)
    _hdr(wr, 3, len(head))
    rfirst = 4
    rlast = rfirst + n - 1 if n else rfirst
    hrng = f"$F${rfirst}:$F${rlast}"
    arng = f"$G${rfirst}:$G${rlast}"
    dv2 = DataValidation(type="list", formula1='"Lulus,Gugur"', allow_blank=True)
    wr.add_data_validation(dv2)
    for i in range(n):
        r = rfirst + i
        wr.cell(row=r, column=1, value=i + 1)
        wr.cell(row=r, column=2, value=f"=Peserta!B{5+i}")
        adm = wr.cell(row=r, column=3, value="Lulus"); dv2.add(adm)          # administrasi: input
        tcol = get_column_letter(pcol0 + i)
        kcol = get_column_letter(kcol0 + i)
        wr.cell(row=r, column=4, value=f"='Evaluasi Teknis'!{tcol}{r_hasil}")
        wr.cell(row=r, column=5, value=f"='Evaluasi Kualifikasi'!{kcol}{rk_hasil}")
        wr.cell(row=r, column=6, value=f"=Peserta!E{5+i}").number_format = _RP
        wr.cell(row=r, column=7,
                value=(f'=IF(AND(C{r}="Lulus",D{r}="LULUS",E{r}="LULUS"),"LULUS","GUGUR")'))
        wr.cell(row=r, column=8,
                value=(f'=IF(G{r}="LULUS",SUMPRODUCT(({arng}="LULUS")*({hrng}<F{r}))+1,"-")'))
        wr.cell(row=r, column=9, value=f'=IF(H{r}=1,"CALON PEMENANG","")')
    pr = rlast + 2
    wr.cell(row=pr, column=2, value="USULAN PEMENANG:").font = Font(bold=True)
    wr.cell(row=pr, column=6, value=(
        f'=IFERROR(INDEX($B${rfirst}:$B${rlast},MATCH(1,$H${rfirst}:$H${rlast},0)),'
        f'"Belum ada yang lulus")')).font = Font(bold=True)
    wr.column_dimensions["B"].width = 28

    fn = f"KertasKerja_{(paket.kode_paket or str(paket.id))}.xlsx".replace("/", "_").replace(" ", "_")
    os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
    wb.save(os.path.join(Config.UPLOAD_FOLDER, fn))
    return True, fn
